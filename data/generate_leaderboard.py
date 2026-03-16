import re
import statistics
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Config
MAX_POINTS = {2019: 380, 2021: 320, 2022: 320, 2023: 320, 2024: 400, 2025: 400}
YEARS = [2019, 2021, 2022, 2023, 2024, 2025]
FILES = {y: f"klassement_{y}.txt" for y in YEARS}

# Parse files
def parse_file(filepath):
    players = []
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.strip()
            # Match: rank. Name Score
            m = re.match(r'^\d+\.\s+(.+?)\s+([\d.]+)\s*$', line)
            if m:
                name = m.group(1).strip()
                score = float(m.group(2))
                players.append((name, score))
    return players

# Normalize names (handle slight variations across years)
NAME_MAP = {
    "Vicky Dhaen": "Vicky D'haen",
    "Vicky D'haen": "Vicky D'haen",
    "Lize Van Assche": "Lize van Assche",
    "Lize Vanassche": "Lize van Assche",
    "Nelly Van Acker": "Nelly Van Acker",
    "Nelly Vanacker": "Nelly Van Acker",
    "Gina Demaertelaere": "Gina De Maertelaere",
    "Gina De Maertelaere": "Gina De Maertelaere",
    "Noa Van de Brande": "Noa Van den Brande",
    "Pieter De Raeymaecker": "Pierre De Raeymaecker",
    "Jef D'haen": "Jef D'haen",
    "Yvan Lammes": "Yvan Lammens",
}

def normalize_name(name):
    return NAME_MAP.get(name, name)

# Load all data
year_data = {}  # {year: [(name, score, rank)]}
all_players = set()

for year in YEARS:
    raw = parse_file(FILES[year])
    parsed = []
    for rank_idx, (name, score) in enumerate(raw, 1):
        name = normalize_name(name)
        parsed.append((name, score, rank_idx))
        all_players.add(name)
    year_data[year] = parsed

# Compute stats per year
year_stats = {}
for year in YEARS:
    scores = [s for _, s, _ in year_data[year]]
    avg = statistics.mean(scores)
    std = statistics.stdev(scores) if len(scores) > 1 else 1
    total = len(scores)
    year_stats[year] = {"mean": avg, "stdev": std, "total": total}

# Build player lookup: {year: {name: (score, rank)}}
player_year = defaultdict(dict)
for year in YEARS:
    for name, score, rank in year_data[year]:
        player_year[year][name] = (score, rank)

# Calculate methods
def z_score(score, year):
    s = year_stats[year]
    return (score - s["mean"]) / s["stdev"]

def percentile_rank(rank, year):
    total = year_stats[year]["total"]
    return 1 - (rank / total)

def points_pct(score, year):
    return score / MAX_POINTS[year]

# Sort players by number of years participated (desc), then alphabetically
sorted_players = sorted(all_players, key=lambda p: (-sum(1 for y in YEARS if p in player_year[y]), p))

# Create Excel workbook
wb = openpyxl.Workbook()

# Colors
BLUE = "005A94"
GOLD = "C9A84C"
DARK = "0A0E14"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
MEDIUM_GRAY = "D9D9D9"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
gold_font = Font(name="Calibri", bold=True, color=DARK, size=11)
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header(ws, row, col, value, fill=None, font=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or header_font
    cell.fill = fill or header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border
    return cell

def style_cell(ws, row, col, value, fmt=None, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = bold_font if bold else normal_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border
    if fmt:
        cell.number_format = fmt
    return cell

# ===================== SHEET 1: Overview =====================
ws = wb.active
ws.title = "Overview"

# Headers
headers = ["#", "Player", "Years"]
for year in YEARS:
    headers.extend([f"{year} Rank", f"{year} Pts"])
headers.extend(["Total Pts", "Avg Pts"])

for col, h in enumerate(headers, 1):
    style_header(ws, 1, col, h)

for row_idx, player in enumerate(sorted_players, 2):
    years_played = sum(1 for y in YEARS if player in player_year[y])
    style_cell(ws, row_idx, 1, row_idx - 1)
    ws.cell(row=row_idx, column=2, value=player).font = bold_font
    ws.cell(row=row_idx, column=2).border = thin_border
    style_cell(ws, row_idx, 3, years_played)

    col = 4
    total_pts = 0
    count = 0
    for year in YEARS:
        if player in player_year[year]:
            score, rank = player_year[year][player]
            style_cell(ws, row_idx, col, rank)
            style_cell(ws, row_idx, col + 1, score, "0.0")
            total_pts += score
            count += 1
        else:
            style_cell(ws, row_idx, col, "-")
            style_cell(ws, row_idx, col + 1, "-")
        col += 2

    style_cell(ws, row_idx, col, total_pts, "0.0", bold=True)
    style_cell(ws, row_idx, col + 1, total_pts / count if count > 0 else 0, "0.0")

# Column widths
ws.column_dimensions["B"].width = 28
for i in range(1, len(headers) + 1):
    if i != 2:
        ws.column_dimensions[get_column_letter(i)].width = 12

# ===================== SHEET 2: Points % =====================
ws2 = wb.create_sheet("Points %")

headers2 = ["#", "Player", "Years"]
for year in YEARS:
    headers2.append(f"{year} %")
headers2.append("Avg %")

for col, h in enumerate(headers2, 1):
    style_header(ws2, 1, col, h)

# Calculate and sort
pct_data = []
for player in sorted_players:
    pcts = []
    for year in YEARS:
        if player in player_year[year]:
            score, _ = player_year[year][player]
            pcts.append(points_pct(score, year))
    avg_pct = statistics.mean(pcts) if pcts else 0
    pct_data.append((player, pcts, avg_pct))

pct_data.sort(key=lambda x: -x[2])

for row_idx, (player, _, avg_pct) in enumerate(pct_data, 2):
    style_cell(ws2, row_idx, 1, row_idx - 1)
    ws2.cell(row=row_idx, column=2, value=player).font = bold_font
    ws2.cell(row=row_idx, column=2).border = thin_border

    col = 3
    for year in YEARS:
        if player in player_year[year]:
            score, _ = player_year[year][player]
            pct = points_pct(score, year)
            style_cell(ws2, row_idx, col, pct, "0.0%")
        else:
            style_cell(ws2, row_idx, col, "-")
        col += 1

    style_cell(ws2, row_idx, col, avg_pct, "0.0%", bold=True)

ws2.column_dimensions["B"].width = 28
for i in range(1, len(headers2) + 1):
    if i != 2:
        ws2.column_dimensions[get_column_letter(i)].width = 12

# ===================== SHEET 3: Z-Score =====================
ws3 = wb.create_sheet("Z-Score")

headers3 = ["#", "Player"]
for year in YEARS:
    headers3.append(f"{year} Z")
headers3.append("Avg Z")

for col, h in enumerate(headers3, 1):
    style_header(ws3, 1, col, h)

z_data = []
for player in sorted_players:
    zscores = []
    for year in YEARS:
        if player in player_year[year]:
            score, _ = player_year[year][player]
            zscores.append(z_score(score, year))
    avg_z = statistics.mean(zscores) if zscores else -999
    z_data.append((player, avg_z))

z_data.sort(key=lambda x: -x[1])

for row_idx, (player, avg_z) in enumerate(z_data, 2):
    style_cell(ws3, row_idx, 1, row_idx - 1)
    ws3.cell(row=row_idx, column=2, value=player).font = bold_font
    ws3.cell(row=row_idx, column=2).border = thin_border

    col = 3
    for year in YEARS:
        if player in player_year[year]:
            score, _ = player_year[year][player]
            z = z_score(score, year)
            style_cell(ws3, row_idx, col, z, "0.00")
        else:
            style_cell(ws3, row_idx, col, "-")
        col += 1

    style_cell(ws3, row_idx, col, avg_z if avg_z > -999 else "-", "0.00" if avg_z > -999 else None, bold=True)

ws3.column_dimensions["B"].width = 28
for i in range(1, len(headers3) + 1):
    if i != 2:
        ws3.column_dimensions[get_column_letter(i)].width = 12

# ===================== SHEET 4: Percentile Rank =====================
ws4 = wb.create_sheet("Percentile Rank")

headers4 = ["#", "Player"]
for year in YEARS:
    headers4.append(f"{year} Pctl")
headers4.append("Avg Pctl")

for col, h in enumerate(headers4, 1):
    style_header(ws4, 1, col, h)

p_data = []
for player in sorted_players:
    pctls = []
    for year in YEARS:
        if player in player_year[year]:
            _, rank = player_year[year][player]
            pctls.append(percentile_rank(rank, year))
    avg_p = statistics.mean(pctls) if pctls else -999
    p_data.append((player, avg_p))

p_data.sort(key=lambda x: -x[1])

for row_idx, (player, avg_p) in enumerate(p_data, 2):
    style_cell(ws4, row_idx, 1, row_idx - 1)
    ws4.cell(row=row_idx, column=2, value=player).font = bold_font
    ws4.cell(row=row_idx, column=2).border = thin_border

    col = 3
    for year in YEARS:
        if player in player_year[year]:
            _, rank = player_year[year][player]
            p = percentile_rank(rank, year)
            style_cell(ws4, row_idx, col, p, "0.0%")
        else:
            style_cell(ws4, row_idx, col, "-")
        col += 1

    style_cell(ws4, row_idx, col, avg_p if avg_p > -999 else "-", "0.0%" if avg_p > -999 else None, bold=True)

ws4.column_dimensions["B"].width = 28
for i in range(1, len(headers4) + 1):
    if i != 2:
        ws4.column_dimensions[get_column_letter(i)].width = 12

# ===================== SHEET 5: Combined Ranking =====================
ws5 = wb.create_sheet("Combined Ranking")

headers5 = ["#", "Player", "Years Played", "Avg Z-Score", "Avg Percentile", "Avg Points %", "Combined Score"]

for col, h in enumerate(headers5, 1):
    style_header(ws5, 1, col, h, gold_fill if col == 7 else None, gold_font if col == 7 else None)

combined = []
for player in sorted_players:
    zscores = []
    pctls = []
    pcts = []
    for year in YEARS:
        if player in player_year[year]:
            score, rank = player_year[year][player]
            zscores.append(z_score(score, year))
            pctls.append(percentile_rank(rank, year))
            pcts.append(points_pct(score, year))

    if zscores:
        avg_z = statistics.mean(zscores)
        avg_p = statistics.mean(pctls)
        avg_pct = statistics.mean(pcts)
        # Combined: average of the 3 normalized methods (Z-score rescaled to 0-1 range roughly)
        # We'll just show all three and rank by Z-score as the "truest" method
        combined.append((player, len(zscores), avg_z, avg_p, avg_pct))

combined.sort(key=lambda x: -x[2])  # Sort by Z-score

for row_idx, (player, years_played, avg_z, avg_p, avg_pct) in enumerate(combined, 2):
    rank = row_idx - 1
    style_cell(ws5, row_idx, 1, rank)
    ws5.cell(row=row_idx, column=2, value=player).font = bold_font
    ws5.cell(row=row_idx, column=2).border = thin_border
    style_cell(ws5, row_idx, 3, years_played)
    style_cell(ws5, row_idx, 4, avg_z, "0.00")
    style_cell(ws5, row_idx, 5, avg_p, "0.0%")
    style_cell(ws5, row_idx, 6, avg_pct, "0.0%")
    # Combined score = weighted average (Z is primary)
    combined_score = (avg_z + 2) / 4  # rough 0-1 normalization
    style_cell(ws5, row_idx, 7, combined_score, "0.00", bold=True)

    # Gold highlight for top 3
    if rank <= 3:
        for c in range(1, 8):
            ws5.cell(row=row_idx, column=c).fill = PatternFill(
                start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"
            )

ws5.column_dimensions["B"].width = 28
for i in range(1, len(headers5) + 1):
    if i != 2:
        ws5.column_dimensions[get_column_letter(i)].width = 16

# ===================== SHEET 6: Year Stats =====================
ws6 = wb.create_sheet("Year Stats")

stat_headers = ["Year", "Max Points", "Players", "Avg Score", "Std Dev", "Min Score", "Max Score", "Winner"]
for col, h in enumerate(stat_headers, 1):
    style_header(ws6, 1, col, h)

for row_idx, year in enumerate(YEARS, 2):
    scores = [s for _, s, _ in year_data[year]]
    winner = year_data[year][0][0]
    style_cell(ws6, row_idx, 1, year, bold=True)
    style_cell(ws6, row_idx, 2, MAX_POINTS[year])
    style_cell(ws6, row_idx, 3, len(scores))
    style_cell(ws6, row_idx, 4, statistics.mean(scores), "0.1")
    style_cell(ws6, row_idx, 5, statistics.stdev(scores), "0.1")
    style_cell(ws6, row_idx, 6, min(scores), "0.1")
    style_cell(ws6, row_idx, 7, max(scores), "0.1")
    ws6.cell(row=row_idx, column=8, value=winner).font = bold_font
    ws6.cell(row=row_idx, column=8).border = thin_border

ws6.column_dimensions["H"].width = 28
for i in range(1, 8):
    ws6.column_dimensions[get_column_letter(i)].width = 14

# Freeze panes on all sheets
for ws_item in [ws, ws2, ws3, ws4, ws5, ws6]:
    ws_item.freeze_panes = "C2"

# Save
output = "fcb_prono_leaderboard.xlsx"
wb.save(output)
print(f"Saved to {output}")
print(f"\nTotal unique players: {len(all_players)}")
print(f"\nTop 10 by Z-Score:")
for i, (player, yrs, avg_z, avg_p, avg_pct) in enumerate(combined[:10], 1):
    print(f"  {i}. {player} (Z={avg_z:.2f}, Pctl={avg_p:.1%}, Pts%={avg_pct:.1%}, {yrs} years)")
